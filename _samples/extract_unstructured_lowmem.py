#!/usr/bin/env python3
"""
extract_unstructured_lowmem.py — streaming, crash-safe, and fast on low RAM

✅ What’s new
- Two progress bars: Files (outer) + per-file (inner, rows/pages).
- Robust streaming readers:
  * PDF: page-by-page via fitz/pdfminer/pypdf
  * TXT: bounded read
  * CSV/TSV: pandas chunksize (no full-file load)
  * Parquet: pyarrow.iter_batches (if available) — otherwise safely skipped
  * Excel (.xlsx): openpyxl read_only streaming (no full sheet in RAM)
- Preflight guard: warns/refuses on suspicious targets (e.g., “VOCAB”) or huge scans without limits.
- Low-memory profile: `--lowmem` one flag to set conservative safe defaults.
- Safety caps: `--max-chars`, `--max-pages`, `--max-rows-per-file`, `--regex-group-size`.
- Targeting: `--include-glob/--exclude-glob`, `--order`, `--max-files`.

CLI example
-----------
python scripts\\extract_unstructured_lowmem.py ^
  --in data\\02_interim\\structured_annotated_v2 ^
  --vocab data\\02_interim\\VOCAB_COMPILED\\vocabulary_enriched.parquet ^
  --out data\\02_interim\\unstructured_lowmem_from_structured ^
  --include-glob "*.csv" --include-glob "*.tsv" --include-glob "*.parquet" --include-glob "*.xlsx" ^
  --order size_asc --max-files 1000 --progress --lowmem
"""
from __future__ import annotations
import argparse, csv, json, os, re, sys, traceback, fnmatch
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Optional, Tuple

# Progress bars
try:
    from tqdm.auto import tqdm  # type: ignore
except Exception:
    tqdm = None

def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def log_append(path: Path, msg: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a", encoding="utf-8") as fh:
        fh.write(msg.rstrip() + "\n")

def batch_iter(it: Iterable, size: int) -> Iterator[List]:
    buf: List = []
    for x in it:
        buf.append(x)
        if len(buf) >= size:
            yield buf; buf = []
    if buf: yield buf

def sentence_split(text: str) -> List[str]:
    chunks = re.split(r'(?<=[\.!\?])\s+|\n+', text)
    return [c.strip() for c in chunks if c and c.strip()]

def safe_relpath(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except Exception:
        return str(path)

# ------------------------- Vocabulary -------------------------
@dataclass
class Vocab:
    alias_to_norm: Dict[str, str]
    alias_to_entity: Dict[str, Optional[str]]
    regex_groups: List[re.Pattern]
    max_terms_per_group: int = 800

def load_vocab(vocab_path: Path, max_terms_per_group: int = 800) -> Vocab:
    import pandas as pd
    ext = vocab_path.suffix.lower()
    if ext == ".parquet":
        try:
            df = pd.read_parquet(vocab_path, engine="pyarrow")
        except Exception:
            df = pd.read_parquet(vocab_path)
    elif ext in {".csv",".tsv"}:
        sep = "," if ext==".csv" else "\t"
        df = pd.read_csv(vocab_path, sep=sep, low_memory=True, on_bad_lines="skip", encoding="utf-8", engine="python")
    elif ext == ".json":
        try: df = pd.read_json(vocab_path, lines=True)
        except ValueError: df = pd.read_json(vocab_path)
    else:
        raise ValueError(f"Unsupported vocab format: {ext}")

    cols = {c.lower(): c for c in df.columns}
    term_cols = [cols[k] for k in ["term","alias","surface","pattern"] if k in cols]
    norm_col  = next((cols[k] for k in ["norm","normalized","canonical","id","label"] if k in cols), None)
    ent_col   = next((cols[k] for k in ["entity","type","class","category"] if k in cols), None)
    if not term_cols:
        raise ValueError(f"No term-like columns in {list(df.columns)}")

    alias_to_norm, alias_to_entity = {}, {}
    for _, row in df.iterrows():
        norm_val = (str(row[norm_col]).strip() if norm_col and pd.notna(row[norm_col]) else "")
        ent_val  = (str(row[ent_col]).strip()  if ent_col  and pd.notna(row[ent_col])  else None)
        for tcol in term_cols:
            if tcol not in row or pd.isna(row[tcol]): continue
            alias = str(row[tcol]).strip()
            if not alias: continue
            key = alias.casefold()
            alias_to_norm[key] = norm_val or alias
            alias_to_entity[key] = ent_val
    if not alias_to_norm:
        raise ValueError("Empty vocab after normalization.")

    terms = list(alias_to_norm.keys())
    terms.sort(key=len, reverse=True)
    regex_groups: List[re.Pattern] = []
    for group in batch_iter(terms, max_terms_per_group):
        pat = r"(?<!\w)(" + "|".join(re.escape(t) for t in group) + r")(?!\w)"
        regex_groups.append(re.compile(pat, flags=re.IGNORECASE))
    return Vocab(alias_to_norm, alias_to_entity, regex_groups, max_terms_per_group)

# ------------------------- Readers (streaming) -------------------------
def read_text_file(path: Path, max_chars: int) -> Iterator[Tuple[int, str]]:
    with open(path, "r", encoding="utf-8", errors="ignore") as fh:
        buf = fh.read(max_chars if max_chars > 0 else None)
        if buf:
            yield 0, buf

def read_delim_chunks(path: Path, sep: str, max_chars: int, chunksize: int, inner_bar=None) -> Iterator[Tuple[int, str]]:
    """CSV/TSV: stream chunks; choose text-like cols on first chunk."""
    import pandas as pd
    try:
        reader = pd.read_csv(path, sep=sep, dtype=str, on_bad_lines="skip",
                             encoding="utf-8", engine="python", chunksize=chunksize)
    except Exception as e:
        raise RuntimeError(f"read_csv failed: {e}")

    text_cols: Optional[List[str]] = None
    chars_left = max_chars if max_chars > 0 else None
    for i, chunk in enumerate(reader):
        if inner_bar: inner_bar.update(len(chunk))
        if text_cols is None:
            # pick text-like columns (avg len > 15)
            text_cols = []
            for c in chunk.columns:
                s = chunk[c].dropna().astype(str)
                if not s.empty and s.str.len().mean() > 15:
                    text_cols.append(c)
        if not text_cols:
            continue
        # flatten chunk text
        sflat = chunk[text_cols].astype(str).fillna("").agg(" ".join, axis=1)
        text = " ".join(sflat.tolist())
        if chars_left is not None:
            if chars_left <= 0: break
            if len(text) > chars_left:
                text = text[:chars_left]
                chars_left = 0
            else:
                chars_left -= len(text)
        if text.strip():
            yield i, text

def read_parquet_batches(path: Path, max_chars: int, batch_rows: int, inner_bar=None) -> Iterator[Tuple[int, str]]:
    """Parquet: pyarrow streaming over string-ish columns. Skips if pyarrow missing."""
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq
    except Exception:
        raise RuntimeError("pyarrow not available; install it or skip Parquet files.")
    pf = pq.ParquetFile(str(path))
    # pick string/large_string columns by schema
    str_cols = [i for i, f in enumerate(pf.schema_arrow) if pa.types.is_string(f.type) or pa.types.is_large_string(f.type)]
    if not str_cols:
        return
    chars_left = max_chars if max_chars > 0 else None
    bi = 0
    for rb in pf.iter_batches(batch_size=max(1024, batch_rows)):
        if inner_bar: inner_bar.update(rb.num_rows)
        cols = [rb.column(i) for i in str_cols]
        # concatenate string columns row-wise
        # convert to pyarrow arrays -> python strings
        rows = []
        nrows = rb.num_rows
        for r in range(nrows):
            parts = []
            for col in cols:
                val = col[r].as_py()
                if val: parts.append(str(val))
            if parts:
                rows.append(" ".join(parts))
        text = " ".join(rows)
        if not text:
            bi += 1
            continue
        if chars_left is not None:
            if chars_left <= 0: break
            if len(text) > chars_left:
                text = text[:chars_left]
                chars_left = 0
            else:
                chars_left -= len(text)
        yield bi, text
        bi += 1

def read_excel_stream(path: Path, max_chars: int, inner_bar=None) -> Iterator[Tuple[int, str]]:
    """Excel .xlsx: openpyxl read_only streaming; choose text-like cols based on first ~1000 rows."""
    try:
        from openpyxl import load_workbook
    except Exception:
        raise RuntimeError("openpyxl not available; install it or convert xlsx to csv.")
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    chars_left = max_chars if max_chars > 0 else None
    block_index = 0
    for sheet in wb.worksheets:
        # pass 1: sample rows to decide text-like columns
        sample_rows = 0
        sums = {}
        counts = {}
        for row in sheet.iter_rows(values_only=True):
            sample_rows += 1
            for idx, val in enumerate(row or []):
                if val is None: continue
                s = str(val)
                sums[idx] = sums.get(idx, 0) + len(s)
                counts[idx] = counts.get(idx, 0) + 1
            if sample_rows >= 1000:
                break
        text_cols = [i for i in sums if counts.get(i,0) and (sums[i]/counts[i]) > 15]
        # pass 2: stream again
        # openpyxl can't rewind easily; reopen the sheet
        sheet2 = wb[sheet.title]
        buf: List[str] = []
        rows_seen = 0
        for row in sheet2.iter_rows(values_only=True):
            rows_seen += 1
            if inner_bar and rows_seen % 500 == 0:
                inner_bar.update(500)
            parts = []
            for idx in text_cols:
                if idx < len(row or []):
                    val = row[idx]
                    if val is not None:
                        parts.append(str(val))
            if parts:
                buf.append(" ".join(parts))
            # flush periodically
            if len(buf) >= 1000:
                text = " ".join(buf); buf = []
                if chars_left is not None:
                    if chars_left <= 0: break
                    if len(text) > chars_left:
                        text = text[:chars_left]; chars_left = 0
                    else:
                        chars_left -= len(text)
                if text:
                    yield block_index, text
                    block_index += 1
        # flush tail
        if buf:
            text = " ".join(buf); buf = []
            if chars_left is not None:
                if chars_left > 0:
                    if len(text) > chars_left:
                        text = text[:chars_left]; chars_left = 0
                    else:
                        chars_left -= len(text)
                else:
                    text = ""
            if text:
                yield block_index, text
                block_index += 1

def read_pdf_pages_auto(path: Path, max_chars: int, backend: str, inner_bar=None, max_pages: int = 0) -> Iterator[Tuple[int, str]]:
    chars_left = max_chars if max_chars > 0 else None
    pages_seen = 0
    def cut(txt: str) -> str:
        nonlocal chars_left
        if chars_left is None: return txt
        if chars_left <= 0: return ""
        if len(txt) > chars_left:
            t = txt[:chars_left]; chars_left = 0; return t
        chars_left -= len(txt); return txt

    def yield_pages(texts: Iterable[Tuple[int,str]]):
        nonlocal pages_seen
        for i, txt in texts:
            if max_pages and pages_seen >= max_pages:
                break
            pages_seen += 1
            if inner_bar: inner_bar.update(1)
            chunk = cut(txt or "")
            if not chunk: break
            yield i, chunk

    tried = []
    if backend in ("auto","fitz"):
        try:
            import fitz
            with fitz.open(str(path)) as doc:
                gen = ((i, page.get_text("text") or "") for i, page in enumerate(doc))
                for tup in yield_pages(gen): yield tup
            return
        except Exception as e:
            tried.append(("fitz", str(e)))
            if backend != "auto": raise
    if backend in ("auto","pdfminer"):
        try:
            from pdfminer.high_level import extract_text
            txt = extract_text(str(path)) or ""
            pages = txt.split("\f") if "\f" in txt else [txt]
            gen = ((i, p) for i, p in enumerate(pages))
            for tup in yield_pages(gen): yield tup
            return
        except Exception as e:
            tried.append(("pdfminer", str(e)))
            if backend != "auto": raise
    if backend in ("auto","pypdf"):
        try:
            try:
                from pypdf import PdfReader
            except Exception:
                from PyPDF2 import PdfReader  # type: ignore
            reader = PdfReader(str(path))
            gen = []
            for i, page in enumerate(reader.pages):
                try: txt = page.extract_text() or ""
                except Exception: txt = ""
                gen.append((i, txt))
            for tup in yield_pages(gen): yield tup
            return
        except Exception as e:
            tried.append(("pypdf", str(e)))
            raise RuntimeError(f"PDF backends failed: {tried}")

# ------------------------- Matching -------------------------
def find_mentions(text: str, vocab: Vocab) -> List[Tuple[int,int,str,str,Optional[str]]]:
    out = []
    for rgx in vocab.regex_groups:
        for m in rgx.finditer(text):
            surface = m.group(1)
            key = surface.casefold()
            out.append((m.start(1), m.end(1), surface,
                        vocab.alias_to_norm.get(key, surface),
                        vocab.alias_to_entity.get(key)))
    out.sort(key=lambda x: (x[0], -(x[1]-x[0])))
    dedup = []
    last = -1
    for s,e,sv,nm,en in out:
        if s >= last:
            dedup.append((s,e,sv,nm,en)); last = e
    return dedup

# ------------------------- Core -------------------------
def process_document(
    file_path: Path, input_root: Path, out_root: Path, vocab: Vocab,
    max_chars: int, context_chars: int,
    emit_sent_text: bool, emit_neighbors: bool,
    pdf_backend: str, max_pages: int, max_rows: int, csv_chunksize: int,
    parquet_batch_rows: int,
    show_inner_bar: bool
) -> Tuple[int,int,int]:
    rel = safe_relpath(file_path, input_root)
    suffix = file_path.suffix.lower()
    ann_path = out_root / "annotations.jsonl"
    pages = sentences = mentions = 0

    # inner progress bar (indeterminate for CSV/Excel; per-page for PDF)
    inner_bar = None
    if show_inner_bar and tqdm is not None:
        unit = "page" if suffix == ".pdf" else ("row" if suffix in {".csv",".tsv",".xlsx",".xls",".parquet"} else "chunk")
        inner_bar = tqdm(total=None, desc=f"{Path(rel).name}", unit=unit, leave=False, dynamic_ncols=True)

    def emit_block(block_index: int, text: str):
        nonlocal pages, sentences, mentions
        pages += 1
        sents = sentence_split(text)
        sentences += len(sents)
        with open(ann_path, "a", encoding="utf-8") as ann_f:
            for si, sent in enumerate(sents):
                hits = find_mentions(sent, vocab); mentions += len(hits)
                prev_sent = sents[si-1] if emit_neighbors and si>0 else None
                next_sent = sents[si+1] if emit_neighbors and si+1<len(sents) else None
                for s,e,surface,norm,entity in hits:
                    left = max(0, s - context_chars); right = min(len(sent), e + context_chars)
                    rec = {
                        "doc_path": rel, "page_index": block_index, "sentence_index": si,
                        "char_start": int(s), "char_end": int(e),
                        "surface": surface, "normalized": norm, "entity": entity,
                        "snippet": sent[left:right],
                    }
                    if emit_sent_text: rec["sentence_text"] = sent
                    if emit_neighbors:
                        rec["prev_sentence"] = prev_sent; rec["next_sentence"] = next_sent
                    ann_f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    try:
        if suffix == ".pdf":
            for bi, text in read_pdf_pages_auto(file_path, max_chars, pdf_backend, inner_bar, max_pages=max_pages):
                if text: emit_block(bi, text)
        elif suffix in {".txt",".md",".rtf",".text"}:
            for bi, text in read_text_file(file_path, max_chars):
                if text: emit_block(bi, text)
        elif suffix in {".csv",".tsv"}:
            sep = "," if suffix == ".csv" else "\t"
            rows_seen = 0
            for bi, text in read_delim_chunks(file_path, sep, max_chars, csv_chunksize, inner_bar):
                if text: emit_block(bi, text)
                rows_seen += csv_chunksize
                if max_rows and rows_seen >= max_rows:
                    break
        elif suffix == ".parquet":
            try:
                rows_seen = 0
                for bi, text in read_parquet_batches(file_path, max_chars, parquet_batch_rows, inner_bar):
                    if text: emit_block(bi, text)
                    rows_seen += parquet_batch_rows
                    if max_rows and rows_seen >= max_rows:
                        break
            except RuntimeError as e:
                raise RuntimeError(f"Parquet skipped ({e})")
        elif suffix in {".xlsx"}:
            rows_seen = 0
            for bi, text in read_excel_stream(file_path, max_chars, inner_bar):
                if text: emit_block(bi, text)
                rows_seen += 1000  # approximate block
                if max_rows and rows_seen >= max_rows:
                    break
        else:
            # other types: skip
            pass
    finally:
        if inner_bar: inner_bar.close()

    return pages, sentences, mentions

# ------------------------- Index -------------------------
def write_doc_index_header(index_csv: Path) -> None:
    if not index_csv.exists():
        ensure_dir(index_csv.parent)
        with open(index_csv, "w", encoding="utf-8", newline="") as fh:
            csv.writer(fh).writerow(["doc_path","ext","size_bytes","pages","sentences","mentions"])

def append_doc_index(index_csv: Path, row: List) -> None:
    with open(index_csv, "a", encoding="utf-8", newline="") as fh:
        csv.writer(fh).writerow(row)

# ------------------------- Preflight -------------------------
def preflight_or_exit(input_root: Path, files: List[Path], args):
    upper_path = str(input_root).upper()
    n = len(files)
    suf = {}
    for p in files[: min(n, 2000)]:
        suf[p.suffix.lower()] = suf.get(p.suffix.lower(), 0) + 1
    suspicious = ("VOCAB" in upper_path and "UNSTRUCTURED" not in upper_path and "STRUCTURED" not in upper_path)
    huge = (n >= args.confirm_threshold and not args.max_files)
    if suspicious or huge:
        print("\n[PRE-FLIGHT] Sanity check")
        print(f"  target: {input_root}")
        print(f"  files found: {n} (sample by suffix: {suf})")
        preview = [str(p.relative_to(input_root)) for p in files[:10]]
        print("  first 10 files:")
        for s in preview: print("    ·", s)
        if suspicious:
            print("  WARN: path contains 'VOCAB' — likely not what you intended.")
        if huge and not args.max_files:
            print(f"  WARN: scanning {n} files without --max-files (threshold={args.confirm_threshold}).")
        if not args.yes:
            print("\n[PRE-FLIGHT] Refusing to proceed. Re-run with:")
            print("  • --yes  (if you’re sure), or")
            print("  • --max-files 500 --order size_asc  (pilot safely)\n")
            sys.exit(2)

# ------------------------- Main -------------------------
def main(argv=None):
    ap = argparse.ArgumentParser(description="Low-memory unstructured extractor (streaming + progress bars)")
    ap.add_argument("--in","--input",  dest="input_dir",  required=True)
    ap.add_argument("--vocab",         dest="vocab_path", required=True)
    ap.add_argument("--out","--output",dest="out_dir",    required=True)

    # Performance & safety
    ap.add_argument("--max-chars",           type=int, default=200_000, help="Max chars per file (cap work)")
    ap.add_argument("--max-pages",           type=int, default=0, help="For PDFs, stop after this many pages (0=all)")
    ap.add_argument("--max-rows-per-file",   type=int, default=0, help="For tables, cap number of rows scanned (0=all)")
    ap.add_argument("--regex-group-size",    type=int, default=400)
    ap.add_argument("--context-chars",       type=int, default=60)
    ap.add_argument("--csv-chunksize",       type=int, default=5000)
    ap.add_argument("--parquet-batch-rows",  type=int, default=8192)

    # Output verbosity
    ap.add_argument("--emit-sent-text",      action="store_true")
    ap.add_argument("--emit-neighbors",      action="store_true")

    # Targeting & ordering
    ap.add_argument("--max-files",           type=int, default=0)
    ap.add_argument("--order",               type=str, default="name", choices=["name","size_asc","size_desc"])
    ap.add_argument("--include-glob",        action="append", default=None, help="Process only files matching this glob (repeatable)")
    ap.add_argument("--exclude-glob",        action="append", default=None, help="Skip files matching this glob (repeatable)")

    # PDF backend
    ap.add_argument("--pdf-backend",         type=str, default="fitz", choices=["auto","fitz","pdfminer","pypdf"])

    # Progress & preflight
    ap.add_argument("--progress",            action="store_true")
    ap.add_argument("--no-progress",         dest="progress", action="store_false")
    ap.set_defaults(progress=True)
    ap.add_argument("--yes",                 action="store_true", help="Skip preflight confirmation for large scans")
    ap.add_argument("--confirm-threshold",   type=int, default=3000, help="Warn/refuse if scanning more than this many files")

    # Low-memory preset
    ap.add_argument("--lowmem",              action="store_true", help="Conservative memory profile")

    args = ap.parse_args(argv)

    input_root = Path(args.input_dir).resolve()
    out_root   = Path(args.out_dir).resolve()
    vocab_path = Path(args.vocab_path).resolve()

    ensure_dir(out_root)
    err_log = out_root / "logs" / "errors.log"
    index_csv = out_root / "doc_index.csv"
    write_doc_index_header(index_csv)

    # keep numeric libs single-threaded
    os.environ.setdefault("OMP_NUM_THREADS","1")
    os.environ.setdefault("MKL_NUM_THREADS","1")
    os.environ.setdefault("OPENBLAS_NUM_THREADS","1")

    # lowmem preset
    if args.lowmem:
        args.max_chars = min(args.max_chars, 150_000)
        args.regex_group_size = min(args.regex_group_size, 300)
        args.context_chars = min(args.context_chars, 40)
        if not getattr(args, "emit_sent_text", False):
            args.emit_sent_text = False
        if not getattr(args, "emit_neighbors", False):
            args.emit_neighbors = False
        # progress bars can add overhead in some shells; we keep them on by default here
        args.csv_chunksize = min(args.csv_chunksize, 4000)
        args.parquet_batch_rows = min(args.parquet_batch_rows, 8192)
        if not args.max_files:
            args.max_files = 1000
        if args.order == "name":
            args.order = "size_asc"

    print(f"[INFO] Loading vocabulary from: {vocab_path}")
    vocab = load_vocab(vocab_path, max_terms_per_group=max(args.regex_group_size, 50))

    print(f"[INFO] Scanning input: {input_root}")
    files: List[Path] = []
    for p in input_root.rglob("*"):
        if not p.is_file(): continue
        if p.suffix.lower() not in {".txt",".md",".rtf",".text",".pdf",".csv",".tsv",".parquet",".xlsx",".xls"}:
            continue
        files.append(p)

    # include/exclude
    if args.include_glob:
        files = [f for f in files if any(fnmatch.fnmatch(str(f).replace("\\","/"), g) for g in args.include_glob)]
    if args.exclude_glob:
        files = [f for f in files if not any(fnmatch.fnmatch(str(f).replace("\\","/"), g) for g in args.exclude_glob)]

    # ordering
    if args.order == "name":
        files.sort(key=lambda p: str(p).lower())
    elif args.order == "size_asc":
        files.sort(key=lambda p: (p.stat().st_size if p.exists() else 0))
    else:
        files.sort(key=lambda p: -(p.stat().st_size if p.exists() else 0))

    total_files = len(files)
    if args.max_files and args.max_files > 0:
        files = files[:args.max_files]
    print(f"[INFO] Found {total_files} files; processing {len(files)}")

    # preflight safety
    preflight_or_exit(input_root, files, args)

    # outer progress bar
    outer_bar = tqdm(total=len(files), desc="Files", unit="file", dynamic_ncols=True) if args.progress and tqdm else None

    total_pages = total_sents = total_mentions = 0
    for i, fp in enumerate(files, 1):
        rel, sz = safe_relpath(fp, input_root), -1
        try:
            sz = fp.stat().st_size if fp.exists() else -1
            print(f"[RUN] {i}/{len(files)} {rel}")
            pages, sents, mentions = process_document(
                fp, input_root, out_root, vocab,
                max_chars=args.max_chars,
                context_chars=args.context_chars,
                emit_sent_text=args.emit_sent_text,
                emit_neighbors=args.emit_neighbors,
                pdf_backend=args.pdf_backend,
                max_pages=args.max_pages,
                max_rows=args.max_rows_per_file,
                csv_chunksize=args.csv_chunksize,
                parquet_batch_rows=args.parquet_batch_rows,
                show_inner_bar=bool(args.progress and tqdm)
            )
            append_doc_index(index_csv, [rel, fp.suffix.lower(), sz, pages, sents, mentions])
            total_pages += pages; total_sents += sents; total_mentions += mentions
            if outer_bar: outer_bar.update(1); outer_bar.set_postfix(pages=pages, mentions=mentions)
        except Exception as e:
            append_doc_index(index_csv, [rel, fp.suffix.lower(), sz, 0, 0, 0])
            log_append(err_log, f"{str(fp)}\t{type(e).__name__}: {e}")
            log_append(err_log, traceback.format_exc())
            if not outer_bar: print(f"[WARN] Failed: {fp} -> {e}")
            else: outer_bar.update(1); outer_bar.set_postfix(error="1")
    if outer_bar: outer_bar.close()
    print(f"[DONE] Docs={len(files)} | Pages={total_pages} | Sentences={total_sents} | Mentions={total_mentions}")
    print(f"[OUT]  {out_root / 'annotations.jsonl'}")
    print(f"[OUT]  {index_csv}")

if __name__ == "__main__":
    main()
