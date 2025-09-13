#!/usr/bin/env python3
"""
link_patients_across_cohorts.py
--------------------------------
Build a MASTER PATIENT INDEX by scanning structured tables for patient/sample IDs.

Design goals
- Low memory: stream CSV/TSV in chunks, Parquet in batches, Excel in read-only rows
- Predictable: always attempt header harvest first (find IDs in column names)
- Robust: anticipate common failures and apply fixes or safe fallbacks automatically
- Deterministic: union multiple IDs per row; stable master ids (sequential or hash)
- Transparent: progress bars, per-file stats, temp JSONL edges, clear logs

Outputs (in --out):
- master_patient_index.parquet (or .csv if parquet writer unavailable)
    columns: [master_patient_id, patient_key]
- id_map.csv  (streamed join from temp edges → master ids)
    columns: [source_file, location, column, raw_value, patient_key, rule, master_patient_id]
- link_report.md (summary)
- logs/errors.log (detailed issues)
- tmp/edges.jsonl (deleted at end unless --keep-temp)

Usage (examples, Windows PowerShell)
------------------------------------
# 1) High-yield metadata files (clinical/sample)
python scripts\\link_patients_across_cohorts.py `
  --in  data\\01_raw\\STRUCTURED `
  --out data\\02_interim\\LINKING_V2 `
  --include-glob "*clinical*" --include-glob "*sample*" --include-glob "*metadata*" --include-glob "*cases*" --include-glob "*patient*" `
  --order size_asc --progress --lowmem --yes

# 2) Whale matrices: harvest headers only (fast, safe)
python scripts\\link_patients_across_cohorts.py `
  --in  data\\01_raw\\STRUCTURED `
  --out data\\02_interim\\LINKING_V2 `
  --include-glob "*rnaseq_all_*.csv" --include-glob "*rnaseq_merged_*.csv" `
  --header-only --order name --progress --yes

# 3) Full pass with config overrides (regex/column hints)
python scripts\\link_patients_across_cohorts.py `
  --in  data\\01_raw\\STRUCTURED `
  --out data\\02_interim\\LINKING_V2 `
  --config config\\dataset_manifest.yml `
  --order size_asc --progress --lowmem --yes
"""
from __future__ import annotations
import argparse, csv, fnmatch, hashlib, io, json, os, re, sys, traceback
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Optional, Set, Tuple

# Optional libs
try:
    from tqdm.auto import tqdm  # type: ignore
except Exception:
    tqdm = None

# YAML config (optional)
try:
    import yaml  # type: ignore
except Exception:
    yaml = None

# --------- Globals & Heuristics ---------
DEFAULT_PATIENT_HINTS = [
    "patient", "patient_id", "case", "case_id", "subject", "subject_id",
    "participant", "participant_id", "donor", "donor_id", "individual",
]
DEFAULT_SAMPLE_HINTS = [
    "sample", "sample_id", "tumor_sample_barcode", "tcga_barcode", "barcode",
    "aliquot", "aliquot_id", "specimen", "specimen_id",
]
DEFAULT_MODEL_HINTS = [
    "model", "model_id", "cell_line", "cell_line_id", "depmap_id", "ach",
]
ID_LIKE_COLS = DEFAULT_PATIENT_HINTS + DEFAULT_SAMPLE_HINTS + DEFAULT_MODEL_HINTS

# TCGA patterns
TCGA_PATIENT_RE = re.compile(r"^(TCGA-[A-Z0-9]{2}-[A-Z0-9]{4})", re.I)
# DepMap / ACH IDs
ACH_RE = re.compile(r"^ACH[-_ ]?\d{4,6}$", re.I)
# UUIDs (GDC cases, etc.)
UUID_RE = re.compile(r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$", re.I)

# --------- Utilities ---------
def ensure_dir(p: Path): p.mkdir(parents=True, exist_ok=True)

def log_append(path: Path, msg: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a", encoding="utf-8") as fh:
        fh.write(msg.rstrip() + "\n")

def norm_token(s: Optional[str]) -> Optional[str]:
    if s is None: return None
    x = re.sub(r"\s+", "", str(s)).upper()
    x = x.replace("_", "-")
    # avoid empty or trivial tokens
    return x if x else None

def is_header_like(cols: Iterable[str]) -> bool:
    # treat "very many" columns as header-heavy (likely sample IDs)
    return len(list(cols)) >= 1000

def candidate_id_cols(columns: Iterable[str], cfg_cols: Dict[str,List[str]]) -> List[str]:
    cols = []
    for c in columns:
        lc = str(c).strip().lower()
        if any(h in lc for h in ID_LIKE_COLS):
            cols.append(c)
        else:
            for group, names in cfg_cols.items():
                if any(n.lower() in lc for n in names):
                    cols.append(c); break
    # stable unique
    out: List[str] = []
    seen = set()
    for c in cols:
        if c not in seen:
            out.append(c); seen.add(c)
    return out

def path_matches(p: Path, include: Optional[List[str]], exclude: Optional[List[str]]) -> bool:
    s = str(p).replace("\\","/")
    if include and not any(fnmatch.fnmatch(s, g) for g in include): return False
    if exclude and any(fnmatch.fnmatch(s, g) for g in exclude): return False
    return True

# --------- ID Derivation Rules ---------
def derive_keys_from_value(raw: str, colname: str) -> List[Tuple[str,str]]:
    """
    Return list of (patient_key, rule_name) derived from a single raw value.
    Multiple keys per value are allowed; caller will union them per row.
    """
    out: List[Tuple[str,str]] = []
    if not raw: return out
    valN = norm_token(raw)
    if not valN: return out

    lc = colname.lower()

    # 1) TCGA patient/sample
    m = TCGA_PATIENT_RE.match(valN)
    if m:
        out.append((m.group(1), "tcga_patient_from_value"))
        # If the raw IS already the 12-char patient, this yields itself.
        # If it's a sample barcode, we still derive the patient core.

    # 2) UUIDs (GDC cases) treated as patient keys as-is
    if UUID_RE.match(valN):
        out.append((valN, "uuid_as_patient"))

    # 3) DepMap / ACH
    if "depmap" in lc or "ach" in lc or ACH_RE.match(valN):
        # normalize to ACH-12345 shape
        ach = valN.replace("ACH", "ACH-").replace("--","-")
        ach = re.sub(r"^ACH-?", "ACH-", ach)
        out.append((ach, "depmap_ach"))

    # 4) Patient-like columns: use value
    if any(h in lc for h in DEFAULT_PATIENT_HINTS):
        out.append((valN, "patient_col_value"))

    # 5) Sample/barcode-like columns: if we didn't catch TCGA above, still use value
    if any(h in lc for h in DEFAULT_SAMPLE_HINTS) and not m:
        out.append((valN, "sample_col_value"))

    # 6) Model-like columns
    if any(h in lc for h in DEFAULT_MODEL_HINTS):
        out.append((valN, "model_col_value"))

    # Deduplicate rules
    seen = set()
    dedup: List[Tuple[str,str]] = []
    for k, rule in out:
        key = (k, rule)
        if key not in seen:
            dedup.append((k, rule)); seen.add(key)
    return dedup

# --------- Union-Find ---------
class UF:
    def __init__(self):
        self.parent: Dict[str,str] = {}
    def find(self, a: str) -> str:
        if a not in self.parent: self.parent[a] = a
        if self.parent[a] != a:
            self.parent[a] = self.find(self.parent[a])
        return self.parent[a]
    def union(self, a: str, b: str):
        ra, rb = self.find(a), self.find(b)
        if ra != rb:
            # simple union by lexical order for determinism
            if ra < rb:
                self.parent[rb] = ra
            else:
                self.parent[ra] = rb

# --------- Readers (streaming, with prediction & fixes) ---------
def read_csv_chunks(fp: Path, usecols: Optional[List[str]], chunksize: int, encodings=("utf-8","latin-1")) -> Iterator[Tuple[Optional[int], 'pd.DataFrame']]:
    import pandas as pd
    for enc in encodings:
        try:
            reader = pd.read_csv(
                fp, dtype=str, usecols=usecols, on_bad_lines="skip",
                encoding=enc, engine="python", chunksize=chunksize, low_memory=True, compression="infer"
            )
            for i, chunk in enumerate(reader):
                yield i, chunk
            return
        except UnicodeDecodeError:
            continue
        except Exception:
            raise
    # if we get here, both encodings failed; let caller handle

def csv_headers(fp: Path) -> List[str]:
    import pandas as pd
    for enc in ("utf-8","latin-1"):
        try:
            df = pd.read_csv(fp, nrows=0, encoding=enc, engine="python", compression="infer")
            return list(df.columns)
        except UnicodeDecodeError:
            continue
        except Exception:
            # fallback: manual first line read
            try:
                with open(fp, "r", encoding=enc, errors="ignore") as fh:
                    line = fh.readline()
                    # rough split (comma/tsv)
                    if "\t" in line: return [c.strip() for c in line.split("\t")]
                    return [c.strip() for c in line.split(",")]
            except Exception:
                pass
    return []

def parquet_headers(fp: Path) -> List[str]:
    try:
        import pyarrow.parquet as pq
    except Exception:
        return []
    try:
        return list(pq.ParquetFile(str(fp)).schema.names)
    except Exception:
        return []

def read_parquet_batches(fp: Path, usecols: Optional[List[str]], batch_rows: int):
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq
    except Exception:
        raise RuntimeError("pyarrow not installed. Run: pip install pyarrow")
    pf = pq.ParquetFile(str(fp))
    # select only string-ish columns if usecols none
    if usecols:
        cols = [c for c in usecols if c in pf.schema.names]
        read_cols = cols if cols else None
    else:
        read_cols = None
    for rb in pf.iter_batches(batch_size=max(1024, batch_rows), columns=read_cols):
        yield rb.to_pandas(types_mapper=str)

def xlsx_headers(fp: Path) -> List[str]:
    try:
        import openpyxl
    except Exception:
        return []
    try:
        wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        # find the first non-empty row as header
        for row in ws.iter_rows(values_only=True):
            vals = [str(x) for x in (row or []) if x is not None]
            if len(vals) >= 2:
                return vals
    except Exception:
        return []
    return []

def read_xlsx_rows(fp: Path, wanted_cols: List[int], row_limit: Optional[int]=None) -> Iterator[List[Optional[str]]]:
    try:
        import openpyxl
    except Exception:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")
    wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    count = 0
    header_done = False
    for row in ws.iter_rows(values_only=True):
        vals = list(row or [])
        if not header_done:
            header_done = True
            continue
        out = []
        for idx in wanted_cols:
            out.append(str(vals[idx]) if idx < len(vals) and vals[idx] is not None else None)
        yield out
        count += 1
        if row_limit and count >= row_limit:
            break

# --------- Preflight ---------
def preflight_or_exit(files: List[Path], yes: bool, threshold: int):
    n = len(files)
    if n >= threshold and not yes:
        print(f"\n[PRE-FLIGHT] You’re about to scan {n} files without confirmation.")
        for p in files[:10]:
            print("  ·", p)
        print("Refusing to proceed. Re-run with --yes or reduce files via --include/--exclude/--max-files.")
        sys.exit(2)

# --------- Main linking logic ---------
def main(argv=None):
    ap = argparse.ArgumentParser(description="Link patients across cohorts (streaming, header-aware, low-memory)")
    ap.add_argument("--in","--input",  dest="input_dir", required=True)
    ap.add_argument("--out",           dest="out_dir",   required=True)

    ap.add_argument("--include-glob", action="append", default=None)
    ap.add_argument("--exclude-glob", action="append", default=None)
    ap.add_argument("--order", choices=["size_asc","name","size_desc"], default="size_asc")
    ap.add_argument("--max-files", type=int, default=0)

    # Performance
    ap.add_argument("--csv-chunksize", type=int, default=50000)
    ap.add_argument("--parquet-batch-rows", type=int, default=8192)
    ap.add_argument("--xlsx-row-limit", type=int, default=250000)

    # Modes
    ap.add_argument("--header-only", action="store_true", help="Harvest IDs from headers only (skip row scans)")
    ap.add_argument("--id-scheme", choices=["sequential","hash"], default="sequential", help="Master ID generation")
    ap.add_argument("--keep-temp", action="store_true", help="Keep temp edges JSONL")

    # Config
    ap.add_argument("--config", type=str, default=None, help="YAML with column hints & regex rules")
    ap.add_argument("--progress", action="store_true")
    ap.add_argument("--no-progress", dest="progress", action="store_false")
    ap.set_defaults(progress=True)
    ap.add_argument("--yes", action="store_true")
    ap.add_argument("--confirm-threshold", type=int, default=2500)
    ap.add_argument("--lowmem", action="store_true")

    args = ap.parse_args(argv)

    # Calm BLAS threads for stability
    os.environ.setdefault("OMP_NUM_THREADS","1")
    os.environ.setdefault("MKL_NUM_THREADS","1")
    os.environ.setdefault("OPENBLAS_NUM_THREADS","1")

    in_root  = Path(args.input_dir).resolve()
    out_root = Path(args.out_dir).resolve()
    ensure_dir(out_root)
    tmp_dir  = out_root / "tmp"
    ensure_dir(tmp_dir)
    edges_jsonl = tmp_dir / "edges.jsonl"

    err_log = out_root / "logs" / "errors.log"
    stats_path = out_root / "stats.json"

    # Config load
    cfg_cols: Dict[str,List[str]] = {"generic": []}
    cfg_regex: List[Tuple[str,str]] = []  # list of (name, regex)
    if args.config:
        if yaml is None:
            print("[WARN] pyyaml not installed; ignoring --config. Run: pip install pyyaml")
        else:
            try:
                with open(args.config, "r", encoding="utf-8") as fh:
                    cfg = yaml.safe_load(fh) or {}
                cfg_cols = cfg.get("column_candidates", {"generic": []}) or {"generic": []}
                # optional regex rules: each item {name: "rule_name", pattern: "regex"}
                for item in cfg.get("regex_rules", []):
                    name = item.get("name","custom")
                    patt = item.get("pattern","")
                    if patt:
                        try:
                            re.compile(patt)
                            cfg_regex.append((name, patt))
                        except re.error as e:
                            log_append(err_log, f"[CONFIG] bad regex {name}: {e}")
            except Exception as e:
                log_append(err_log, f"[CONFIG] load failed: {e}")

    # Discover files
    files: List[Path] = []
    for p in in_root.rglob("*"):
        if not p.is_file(): continue
        ext = p.suffix.lower()
        if ext in {".csv",".tsv",".parquet",".xlsx"}:
            if path_matches(p, args.include_glob, args.exclude_glob):
                files.append(p)

    # Order & cap
    if args.order == "name":
        files.sort(key=lambda p: str(p).lower())
    elif args.order == "size_asc":
        files.sort(key=lambda p: p.stat().st_size if p.exists() else 0)
    else:
        files.sort(key=lambda p: -(p.stat().st_size if p.exists() else 0))
    if args.max_files and args.max_files > 0:
        files = files[:args.max_files]

    # Preflight
    preflight_or_exit(files, yes=args.yes, threshold=args.confirm_threshold)

    # Lowmem tweaks
    if args.lowmem:
        args.csv_chunksize = min(args.csv_chunksize, 40000)
        args.parquet_batch_rows = min(args.parquet_batch_rows, 8192)
        args.xlsx_row_limit = min(args.xlsx_row_limit, 200000)

    # Progress
    outer = tqdm(total=len(files), desc="Files", unit="file", dynamic_ncols=True) if args.progress and tqdm else None

    uf = UF()
    unique_keys: Set[str] = set()
    total_edges = 0
    files_harvested = 0

    def write_edge(file: Path, location: str, column: Optional[str], raw: str, pkey: str, rule: str):
        nonlocal total_edges, unique_keys
        with open(edges_jsonl, "a", encoding="utf-8") as fh:
            fh.write(json.dumps({
                "source_file": str(file),
                "location": location,     # HEADER or ROW
                "column": column or "",
                "raw_value": raw,
                "patient_key": pkey,
                "rule": rule
            }, ensure_ascii=False) + "\n")
        total_edges += 1
        unique_keys.add(pkey)
        uf.union(pkey, pkey)

    # Custom regex rule applier (optional)
    def apply_custom_regexes(val: str) -> List[Tuple[str,str]]:
        out: List[Tuple[str,str]] = []
        if not val or not cfg_regex:
            return out
        for name, patt in cfg_regex:
            try:
                m = re.search(patt, val)
                if m:
                    token = norm_token(m.group(1) if m.groups() else m.group(0))
                    if token:
                        out.append((token, f"regex:{name}"))
            except Exception:
                continue
        return out

    # Process each file
    import pandas as pd  # used in multiple code paths; import once

    for fp in files:
        harvested = 0
        try:
            ext = fp.suffix.lower()

            # 1) HEADER HARVEST (cheap constant mem)
            headers: List[str] = []
            if ext == ".csv" or ext == ".tsv":
                headers = csv_headers(fp)
            elif ext == ".parquet":
                headers = parquet_headers(fp)
            elif ext == ".xlsx":
                headers = xlsx_headers(fp)

            # derive patient keys from header tokens (big matrices)
            if headers:
                # if there are "too many" columns, treat as header-heavy
                if is_header_like(headers) or args.header_only:
                    for h in headers:
                        for k, rule in derive_keys_from_value(h, "HEADER"):
                            write_edge(fp, "HEADER", "HEADER", h, k, rule)
                            harvested += 1

            if args.header_only:
                files_harvested += 1
                if outer: outer.update(1); outer.set_postfix(items=harvested)
                continue  # skip row scans by request

            # 2) ROW SCANS (streaming) — only ID candidate columns
            cand_cols = candidate_id_cols(headers, cfg_cols)
            # Parquet path: schema known; CSV/TSV path: headers are enough; XLSX: map indices

            if ext in {".csv",".tsv"}:
                sep = "," if ext == ".csv" else "\t"
                usecols = cand_cols if cand_cols else None
                inner = tqdm(total=None, desc=fp.name, unit="rows", leave=False, dynamic_ncols=True) if args.progress and tqdm else None
                for i, chunk in read_csv_chunks(fp, usecols, args.csv_chunksize):
                    # If no candidate columns found, pick texty columns heuristically
                    if usecols is None:
                        # heuristic: any col with avg len >= 6
                        usecols = []
                        for c in chunk.columns:
                            s = chunk[c].dropna().astype(str)
                            if not s.empty and s.str.len().mean() >= 6:
                                usecols.append(c)
                        if not usecols:
                            break
                        # keep only those columns
                        chunk = chunk[usecols]
                    # Process rows
                    for _, row in chunk.iterrows():
                        row_keys: List[Tuple[str,str]] = []
                        for c in row.index:
                            raw = row[c]
                            if pd.isna(raw) or raw is None: continue
                            raw = str(raw).strip()
                            if not raw: continue
                            # built-in rules
                            row_keys.extend(derive_keys_from_value(raw, str(c)))
                            # custom regex
                            row_keys.extend(apply_custom_regexes(raw))
                        # union all derived keys for this row; write edges
                        seen_row_keys: Set[str] = set()
                        for k, rule in row_keys:
                            if k not in seen_row_keys:
                                write_edge(fp, "ROW", None, raw=raw, pkey=k, rule=rule)
                                seen_row_keys.add(k)
                                harvested += 1
                        # union within-row keys (co-referent)
                        if len(seen_row_keys) >= 2:
                            ks = list(seen_row_keys)
                            root = ks[0]
                            for kk in ks[1:]:
                                uf.union(root, kk)
                    if inner: inner.update(len(chunk))
                if inner: inner.close()

            elif ext == ".parquet":
                inner = tqdm(total=None, desc=fp.name, unit="rows", leave=False, dynamic_ncols=True) if args.progress and tqdm else None
                try:
                    for chunk in read_parquet_batches(fp, cand_cols if cand_cols else None, args.parquet_batch_rows):
                        if chunk is None or chunk.empty: continue
                        usecols = cand_cols if cand_cols else None
                        if usecols is None:
                            # heuristic pick
                            usecols = []
                            for c in chunk.columns:
                                s = chunk[c].dropna().astype(str)
                                if not s.empty and s.str.len().mean() >= 6:
                                    usecols.append(c)
                            if not usecols:
                                continue
                            chunk = chunk[usecols]
                        for _, row in chunk.iterrows():
                            row_keys: List[Tuple[str,str]] = []
                            for c in row.index:
                                raw = row[c]
                                if pd.isna(raw) or raw is None: continue
                                raw = str(raw).strip()
                                if not raw: continue
                                row_keys.extend(derive_keys_from_value(raw, str(c)))
                                row_keys.extend(apply_custom_regexes(raw))
                            seen_row_keys: Set[str] = set()
                            for k, rule in row_keys:
                                if k not in seen_row_keys:
                                    write_edge(fp, "ROW", None, raw=raw, pkey=k, rule=rule)
                                    seen_row_keys.add(k); harvested += 1
                            if len(seen_row_keys) >= 2:
                                ks = list(seen_row_keys)
                                root = ks[0]
                                for kk in ks[1:]:
                                    uf.union(root, kk)
                        if inner: inner.update(len(chunk))
                except RuntimeError as e:
                    log_append(err_log, f"{fp}\tParquet skipped: {e}")
                if inner: inner.close()

            elif ext == ".xlsx":
                # Map header indices to candidates; fall back to all columns if none
                hdr = headers or []
                wanted_idx: List[int] = []
                if hdr:
                    for i, name in enumerate(hdr):
                        lc = str(name).lower()
                        if any(h in lc for h in ID_LIKE_COLS):
                            wanted_idx.append(i)
                # if none, take first 12 columns as a precaution (keeps memory small)
                if not wanted_idx:
                    wanted_idx = list(range(min(12, len(hdr)))) if hdr else list(range(12))
                inner = tqdm(total=None, desc=fp.name, unit="rows", leave=False, dynamic_ncols=True) if args.progress and tqdm else None
                try:
                    row_iter = read_xlsx_rows(fp, wanted_idx, row_limit=args.xlsx_row_limit)
                    for rowvals in row_iter:
                        row_keys: List[Tuple[str,str]] = []
                        for j, raw in enumerate(rowvals):
                            if not raw: continue
                            row_keys.extend(derive_keys_from_value(raw, hdr[wanted_idx[j]] if j < len(wanted_idx) and wanted_idx[j] < len(hdr) else f"col{j}"))
                            row_keys.extend(apply_custom_regexes(raw))
                        seen_row_keys: Set[str] = set()
                        for k, rule in row_keys:
                            if k not in seen_row_keys:
                                write_edge(fp, "ROW", None, raw=raw, pkey=k, rule=rule)
                                seen_row_keys.add(k); harvested += 1
                        if len(seen_row_keys) >= 2:
                            ks = list(seen_row_keys)
                            root = ks[0]
                            for kk in ks[1:]:
                                uf.union(root, kk)
                        if inner: inner.update(1)
                except RuntimeError as e:
                    log_append(err_log, f"{fp}\tXLSX skipped: {e}")
                if inner: inner.close()

            files_harvested += 1
        except Exception as e:
            log_append(err_log, f"{fp}\t{type(e).__name__}: {e}")
            log_append(err_log, traceback.format_exc())
        finally:
            if outer:
                outer.update(1); outer.set_postfix(items=harvested)

    if outer: outer.close()

    # --------- Build master ids deterministically ---------
    # Collapse to roots
    roots: Dict[str, Set[str]] = {}
    for k in unique_keys:
        r = uf.find(k)
        roots.setdefault(r, set()).add(k)

    # Deterministic ordering of roots
    root_list = sorted(roots.keys())
    # ID assignment
    if args.id_scheme == "hash":
        def mkid(key: str) -> str:
            return "P" + hashlib.sha1(key.encode("utf-8")).hexdigest()[:12].upper()
        root_to_mid = {r: mkid(min(roots[r])) for r in root_list}
    else:
        root_to_mid = {r: f"P{idx:08d}" for idx, r in enumerate(root_list, start=1)}

    # Write master index
    ensure_dir(out_root)
    idx_rows = []
    for r in root_list:
        mid = root_to_mid[r]
        for k in sorted(roots[r]):
            idx_rows.append({"master_patient_id": mid, "patient_key": k})

    idx_path = out_root / "master_patient_index.parquet"
    try:
        import pandas as pd
        pd.DataFrame(idx_rows).to_parquet(idx_path, index=False)
    except Exception:
        # fallback to CSV
        idx_csv = str(idx_path).replace(".parquet",".csv")
        with open(idx_csv, "w", encoding="utf-8", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["master_patient_id","patient_key"])
            for r in root_list:
                mid = root_to_mid[r]
                for k in sorted(roots[r]):
                    w.writerow([mid, k])
        idx_path = Path(idx_csv)

    # Stream out id_map.csv by joining edges.jsonl → master ids
    id_map_path = out_root / "id_map.csv"
    with open(id_map_path, "w", encoding="utf-8", newline="") as fh_out:
        w = csv.writer(fh_out)
        w.writerow(["source_file","location","column","raw_value","patient_key","rule","master_patient_id"])
        with open(edges_jsonl, "r", encoding="utf-8") as fh_in:
            for line in fh_in:
                rec = json.loads(line)
                key = rec["patient_key"]
                mid = root_to_mid[uf.find(key)]
                w.writerow([
                    rec["source_file"], rec["location"], rec["column"],
                    rec["raw_value"], key, rec["rule"], mid
                ])

    # Report & stats
    with open(out_root / "link_report.md", "w", encoding="utf-8") as fh:
        fh.write("# Patient Link Report\n\n")
        fh.write(f"- Files processed: {len(files)} (harvested={files_harvested})\n")
        fh.write(f"- Edges observed: {total_edges}\n")
        fh.write(f"- Unique patient keys: {len(unique_keys)}\n")
        fh.write(f"- Master patients: {len(root_list)}\n")
        fh.write("\n## Sample master ids\n")
        for r in root_list[:10]:
            mid = root_to_mid[r]
            sample = ", ".join(sorted(list(roots[r]))[:6])
            fh.write(f"- {mid}: {sample}\n")

    stats = {
        "files": len(files),
        "files_harvested": files_harvested,
        "edges": total_edges,
        "unique_keys": len(unique_keys),
        "masters": len(root_list),
        "id_scheme": args.id_scheme
    }
    with open(stats_path, "w", encoding="utf-8") as fh:
        json.dump(stats, fh, indent=2)

    # Cleanup
    if not args.keep_temp:
        try: edges_jsonl.unlink()
        except Exception: pass

    print(f"[OUT] {idx_path}")
    print(f"[OUT] {id_map_path}")
    print(f"[OUT] {out_root / 'link_report.md'}")

if __name__ == "__main__":
    main()
