#!/usr/bin/env python3
# stream_extract_and_discover.py — streaming extractor + vocab matcher + unknown-term discovery (≤700 MB)
from __future__ import annotations
import argparse, csv, json, os, re, sys, traceback, fnmatch, collections
from pathlib import Path
from typing import List, Tuple, Optional, Iterable, Dict

try:
    from tqdm.auto import tqdm  # optional nice progress bars
except Exception:
    tqdm = None

# ---------- small utils ----------
def ensure_dir(p: Path): p.mkdir(parents=True, exist_ok=True)
def safe_rel(root: Path, p: Path) -> str:
    try: return str(p.relative_to(root))
    except Exception: return str(p)
def sentence_split(text: str) -> List[str]:
    return [s.strip() for s in re.split(r'(?<=[.!?])\s+|\n+', text) if s and s.strip()]

# ---------- vocab ----------
class Vocab:
    __slots__ = ("a2n","a2e","groups")
    def __init__(self, a2n: Dict[str,str], a2e: Dict[str,Optional[str]], groups: List[re.Pattern]):
        self.a2n, self.a2e, self.groups = a2n, a2e, groups

def load_vocab(vpath: Path, max_terms_per_group: int = 400) -> Vocab:
    import pandas as pd
    ext = vpath.suffix.lower()
    if ext == ".parquet":
        try: df = pd.read_parquet(vpath, engine="pyarrow")
        except Exception: df = pd.read_parquet(vpath)
    elif ext in (".csv",".tsv"):
        df = pd.read_csv(vpath, sep="," if ext==".csv" else "\t",
                         low_memory=True, on_bad_lines="skip", encoding="utf-8", engine="python")
    elif ext == ".json":
        try: df = pd.read_json(vpath, lines=True)
        except ValueError: df = pd.read_json(vpath)
    else:
        raise ValueError(f"Unsupported vocab: {ext}")

    cols = {c.lower(): c for c in df.columns}
    term_cols = [cols[k] for k in ("term","alias","surface","pattern") if k in cols]
    norm_col  = next((cols[k] for k in ("norm","normalized","canonical","id","label") if k in cols), None)
    ent_col   = next((cols[k] for k in ("entity","type","class","category") if k in cols), None)
    if not term_cols: raise ValueError("Vocab missing term-like columns")

    a2n: Dict[str,str] = {}; a2e: Dict[str,Optional[str]] = {}
    for _, r in df.iterrows():
        n = (str(r[norm_col]).strip() if norm_col is not None and pd.notna(r[norm_col]) else "")
        e = (str(r[ent_col]).strip()  if ent_col  is not None and pd.notna(r[ent_col])  else None)
        for tc in term_cols:
            v = r.get(tc, None)
            if pd.isna(v): continue
            alias = str(v).strip()
            if not alias: continue
            key = alias.casefold()
            a2n[key] = n or alias
            a2e[key] = e
    if not a2n: raise ValueError("Empty vocab after normalization")

    terms = sorted(a2n.keys(), key=len, reverse=True)
    groups: List[re.Pattern] = []
    step = max(50, max_terms_per_group)
    for i in range(0, len(terms), step):
        pat = r"(?<!\w)(" + "|".join(re.escape(t) for t in terms[i:i+step]) + r")(?!\w)"
        groups.append(re.compile(pat, flags=re.IGNORECASE))
    return Vocab(a2n, a2e, groups)

# ---------- unknown discovery ----------
STOPWORDS = {
    "the","and","for","that","with","this","from","were","have","has","had","but","are","was","be","by","not",
    "into","than","then","over","under","onto","also","such","can","could","should","would","may","might","our",
    "their","there","where","which","while","within","between","each","per","via","using","used","based"
}
TOKEN_RE = re.compile(r"[A-Za-z][A-Za-z0-9\-+/_.]{4,}")  # ≥5 chars, bio/chem friendly
def find_unknown_tokens(text: str, vocab_keys:set[str], min_len:int) -> list[str]:
    out = []
    for tok in TOKEN_RE.findall(text):
        k = tok.casefold()
        if len(tok) >= min_len and k not in vocab_keys and k not in STOPWORDS:
            out.append(k)
    return out

# ---------- streaming readers ----------
def read_text(path: Path, max_chars: int) -> Iterable[Tuple[int,str]]:
    with open(path, "r", encoding="utf-8", errors="ignore") as fh:
        chunk = fh.read(max_chars if max_chars>0 else None)
        if chunk: yield 0, chunk

def read_csv_tsv(path: Path, sep: str, max_chars: int, chunksize: int, pbar=None) -> Iterable[Tuple[int,str]]:
    import pandas as pd
    reader = pd.read_csv(path, sep=sep, dtype=str, on_bad_lines="skip",
                         encoding="utf-8", engine="python", chunksize=chunksize)
    text_cols: Optional[List[str]] = None
    remain = max_chars if max_chars>0 else None
    for i, chunk in enumerate(reader):
        if pbar: pbar.update(len(chunk))
        if text_cols is None:
            text_cols = []
            for c in chunk.columns:
                s = chunk[c].dropna().astype(str)
                if not s.empty and s.str.len().mean() > 15:
                    text_cols.append(c)
        if not text_cols: continue
        s = chunk[text_cols].astype(str).fillna("").agg(" ".join, axis=1)
        text = " ".join(s.tolist())
        if remain is not None:
            if remain <= 0: break
            if len(text) > remain: text, remain = text[:remain], 0
            else: remain -= len(text)
        if text.strip(): yield i, text

def read_parquet(path: Path, max_chars: int, batch_rows: int, pbar=None) -> Iterable[Tuple[int,str]]:
    """Vectorized fast-path using pyarrow.compute; safe fallback if unavailable."""
    try:
        import pyarrow as pa, pyarrow.parquet as pq, pyarrow.compute as pc
        pf = pq.ParquetFile(str(path))
        schema = pf.schema_arrow
        str_idx = [i for i,f in enumerate(schema) if pa.types.is_string(f.type) or pa.types.is_large_string(f.type)]
        if not str_idx: return
        remain = max_chars if max_chars>0 else None
        b = 0
        for rb in pf.iter_batches(batch_size=max(2048, batch_rows)):
            if pbar: pbar.update(rb.num_rows)
            joined = rb.column(str_idx[0]).cast(pa.large_string())
            for ci in str_idx[1:]:
                joined = pc.binary_join_element_wise(joined, pa.scalar(" "), rb.column(ci).cast(pa.large_string()))
            step = 5000
            for s in range(0, len(joined), step):
                seg = joined.slice(s, min(step, len(joined)-s))
                text = " ".join([x for x in seg.to_pylist() if x])
                if not text: continue
                if remain is not None:
                    if remain <= 0: return
                    if len(text) > remain: text, remain = text[:remain], 0
                    else: remain -= len(text)
                yield b, text; b += 1
        return
    except Exception:
        try:
            import pyarrow.parquet as pq
            pf = pq.ParquetFile(str(path))
            remain = max_chars if max_chars>0 else None
            b = 0
            for rb in pf.iter_batches(batch_size=max(1024, batch_rows)):
                if pbar: pbar.update(rb.num_rows)
                tbl = rb.to_pandas()
                str_cols = [c for c in tbl.columns if tbl[c].dtype == object]
                if not str_cols: continue
                text = " ".join(tbl[str_cols].astype(str).fillna("").agg(" ".join, axis=1).tolist())
                if remain is not None:
                    if remain <= 0: return
                    if len(text) > remain: text, remain = text[:remain], 0
                    else: remain -= len(text)
                if text.strip(): yield b, text; b += 1
        except Exception as e:
            raise RuntimeError(f"Parquet skipped: {e}")

def read_xlsx_stream(path: Path, max_chars: int, pbar=None) -> Iterable[Tuple[int,str]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        raise RuntimeError("openpyxl not installed; convert XLSX to CSV")
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    remain = max_chars if max_chars>0 else None; blk = 0
    for ws in wb.worksheets:
        sums, cnt = {}, {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            for j, v in enumerate(row or []):
                if v is None: continue
                s = str(v); sums[j] = sums.get(j,0)+len(s); cnt[j] = cnt.get(j,0)+1
            if i >= 999: break
        txt_cols = [j for j in sums if cnt.get(j,0) and (sums[j]/cnt[j])>15]
        buf: List[str] = []; r = 0
        for row in ws.iter_rows(values_only=True):
            r += 1
            if pbar and r % 500 == 0: pbar.update(500)
            parts = [str(row[j]) for j in txt_cols if j < len(row or []) and row[j] is not None]
            if parts: buf.append(" ".join(parts))
            if len(buf) >= 1000:
                text = " ".join(buf); buf.clear()
                if remain is not None:
                    if remain <= 0: break
                    if len(text) > remain: text, remain = text[:remain], 0
                    else: remain -= len(text)
                if text: yield blk, text; blk += 1
        if buf and (remain is None or remain>0):
            text = " ".join(buf)
            if remain is not None: text = text[:remain]; remain = max(0, remain - len(text))
            if text: yield blk, text; blk += 1

def read_xls_stream(path: Path, max_chars: int, pbar=None) -> Iterable[Tuple[int,str]]:
    try:
        import xlrd  # pip install xlrd==1.2.0 for legacy .xls
    except Exception:
        raise RuntimeError("xlrd not available; convert XLS to CSV")
    wb = xlrd.open_workbook(str(path))
    remain = max_chars if max_chars>0 else None; blk = 0
    for sh in wb.sheets():
        rows, cols = sh.nrows, sh.ncols
        lens = [0]*cols; cnt = [0]*cols
        for i in range(min(rows, 1000)):
            for j in range(cols):
                v = sh.cell_value(i,j)
                if v == "" or v is None: continue
                s = str(v); lens[j]+=len(s); cnt[j]+=1
        txt_cols = [j for j in range(cols) if cnt[j] and (lens[j]/cnt[j])>15]
        buf: List[str] = []
        for i in range(rows):
            if pbar and i % 500 == 0: pbar.update(500)
            parts = []
            for j in txt_cols:
                v = sh.cell_value(i,j)
                if v != "" and v is not None: parts.append(str(v))
            if parts: buf.append(" ".join(parts))
            if len(buf) >= 1000:
                text = " ".join(buf); buf.clear()
                if remain is not None:
                    if remain <= 0: break
                    if len(text) > remain: text, remain = text[:remain], 0
                    else: remain -= len(text)
                if text: yield blk, text; blk += 1
        if buf and (remain is None or remain>0):
            text = " ".join(buf)
            if remain is not None: text = text[:remain]; remain = max(0, remain - len(text))
            if text: yield blk, text; blk += 1

def read_pdf(path: Path, max_chars: int, backend: str, pbar=None, max_pages: int=0) -> Iterable[Tuple[int,str]]:
    remain = max_chars if max_chars>0 else None
    def cut(s: str) -> str:
        nonlocal remain
        if remain is None: return s
        if remain <= 0: return ""
        if len(s) > remain: out, remain = s[:remain], 0; return out
        remain -= len(s); return s
    tried = []
    if backend in ("auto","fitz"):
        try:
            import fitz
            with fitz.open(str(path)) as doc:
                for i, pg in enumerate(doc):
                    if max_pages and i >= max_pages: break
                    if pbar: pbar.update(1)
                    yield i, cut(pg.get_text("text") or "")
            return
        except Exception as e:
            tried.append(("fitz", str(e)))
            if backend != "auto": raise
    if backend in ("auto","pdfminer"):
        try:
            from pdfminer.high_level import extract_text
            txt = extract_text(str(path)) or ""
            pages = txt.split("\f") if "\f" in txt else [txt]
            for i, p in enumerate(pages):
                if max_pages and i >= max_pages: break
                if pbar: pbar.update(1)
                yield i, cut(p)
            return
        except Exception as e:
            tried.append(("pdfminer", str(e)))
            if backend != "auto": raise
    if backend in ("auto","pypdf"):
        try:
            try: from pypdf import PdfReader
            except Exception: from PyPDF2 import PdfReader
            rdr = PdfReader(str(path))
            for i, page in enumerate(rdr.pages):
                if max_pages and i >= max_pages: break
                if pbar: pbar.update(1)
                try: txt = page.extract_text() or ""
                except Exception: txt = ""
                yield i, cut(txt)
            return
        except Exception as e:
            tried.append(("pypdf", str(e)))
            raise RuntimeError(f"PDF backends failed: {tried}")

# ---------- matching ----------
def find_mentions(text: str, vocab: Vocab) -> List[Tuple[int,int,str,str,Optional[str]]]:
    out = []
    for rgx in vocab.groups:
        for m in rgx.finditer(text):
            s,e = m.start(1), m.end(1)
            sv = m.group(1); key = sv.casefold()
            out.append((s,e,sv, vocab.a2n.get(key, sv), vocab.a2e.get(key)))
    out.sort(key=lambda x: (x[0], -(x[1]-x[0])))
    dedup, last = [], -1
    for s,e,sv,nm,en in out:
        if s >= last:
            dedup.append((s,e,sv,nm,en)); last = e
    return dedup

# ---------- per-file processing ----------
def process_file(fp: Path, root: Path, out_dir: Path, vocab: Vocab, vocab_keys:set[str],
                 max_chars:int, ctx:int, emit_sent:bool, emit_neighbors:bool,
                 pdf_backend:str, max_pages:int, max_rows:int,
                 csv_chunksize:int, parquet_batch_rows:int,
                 progress_inner:bool, output_mode:str,
                 discover_unknown:bool, min_unknown_len:int,
                 unk_counts_path:Path) -> Tuple[int,int,int]:
    rel = safe_rel(root, fp); ext = fp.suffix.lower()
    if output_mode == "single":
        tgt = out_dir / "annotations.jsonl"
    else:
        shard = (rel.replace("\\","__").replace("/","__") + ".jsonl")
        tgt = out_dir / "docs" / shard
        ensure_dir(tgt.parent)

    unk_counter = collections.Counter() if discover_unknown else None

    def emit(block_idx: int, text: str):
        nonlocal pages, sentences, mentions
        pages += 1
        sents = sentence_split(text)
        sentences += len(sents)
        with open(tgt, "a", encoding="utf-8") as fh:
            for i, sent in enumerate(sents):
                hits = find_mentions(sent, vocab); mentions += len(hits)
                if unk_counter is not None:
                    for u in find_unknown_tokens(sent, vocab_keys, min_unknown_len):
                        unk_counter[u] += 1
                Ls = sents[i-1] if emit_neighbors and i>0 else None
                Ns = sents[i+1] if emit_neighbors and i+1<len(sents) else None
                for s,e,sv,nm,en in hits:
                    L = max(0, s - ctx); R = min(len(sent), e + ctx)
                    rec = {"doc_path": rel, "page_index": block_idx, "sentence_index": i,
                           "char_start": int(s), "char_end": int(e),
                           "surface": sv, "normalized": nm, "entity": en,
                           "snippet": sent[L:R]}
                    if emit_sent: rec["sentence_text"] = sent
                    if emit_neighbors: rec["prev_sentence"] = Ls; rec["next_sentence"] = Ns
                    fh.write(json.dumps(rec, ensure_ascii=False) + "\n")

    unit = "page" if ext==".pdf" else ("row" if ext in {".csv",".tsv",".xlsx",".xls",".parquet"} else "chunk")
    pbar = tqdm(total=None, desc=Path(rel).name, unit=unit, leave=False, dynamic_ncols=True) if (progress_inner and tqdm) else None

    pages = sentences = mentions = 0
    try:
        if ext == ".pdf":
            for bi, t in read_pdf(fp, max_chars, pdf_backend, pbar, max_pages=max_pages):
                if t: emit(bi, t)
        elif ext in {".txt",".md",".rtf",".text"}:
            for bi, t in read_text(fp, max_chars):
                if t: emit(bi, t)
        elif ext in {".csv",".tsv"}:
            sep = "," if ext==".csv" else "\t"
            seen = 0
            for bi, t in read_csv_tsv(fp, sep, max_chars, csv_chunksize, pbar):
                if t: emit(bi, t)
                seen += csv_chunksize
                if max_rows and seen >= max_rows: break
        elif ext == ".parquet":
            seen = 0
            for bi, t in read_parquet(fp, max_chars, parquet_batch_rows, pbar):
                if t: emit(bi, t)
                seen += parquet_batch_rows
                if max_rows and seen >= max_rows: break
        elif ext == ".xlsx":
            seen = 0
            for bi, t in read_xlsx_stream(fp, max_chars, pbar):
                if t: emit(bi, t)
                seen += 1000
                if max_rows and seen >= max_rows: break
        elif ext == ".xls":
            seen = 0
            for bi, t in read_xls_stream(fp, max_chars, pbar):
                if t: emit(bi, t)
                seen += 1000
                if max_rows and seen >= max_rows: break
    finally:
        if pbar: pbar.close()

    if unk_counter:
        with open(unk_counts_path, "a", encoding="utf-8") as fh:
            fh.write(json.dumps({"doc_path": rel, "counts": unk_counter.most_common(2000)}, ensure_ascii=False) + "\n")
    return pages, sentences, mentions

# ---------- main ----------
def main(argv=None):
    ap = argparse.ArgumentParser(description="Streaming extractor (PDF/CSV/TSV/Parquet/Excel/TXT) with vocab matching + unknown discovery")
    ap.add_argument("--input","--in",  dest="input_path", required=True, help="Folder OR single file")
    ap.add_argument("--vocab",         dest="vocab_path", required=True)
    ap.add_argument("--output","--out",dest="out_dir",    required=True)

    # performance & safety
    ap.add_argument("--max-chars", type=int, default=200_000)
    ap.add_argument("--max-pages", type=int, default=0)
    ap.add_argument("--max-rows-per-file", type=int, default=0)
    ap.add_argument("--csv-chunksize", type=int, default=4000)
    ap.add_argument("--parquet-batch-rows", type=int, default=8192)

    # verbosity / context
    ap.add_argument("--context-chars", type=int, default=60)
    ap.add_argument("--emit-sent-text", action="store_true")
    ap.add_argument("--emit-neighbors", action="store_true")

    # targeting & ordering
    ap.add_argument("--max-files", type=int, default=0)
    ap.add_argument("--order", choices=["name","size_asc","size_desc"], default="size_asc")
    ap.add_argument("--include-glob", action="append", default=None)
    ap.add_argument("--exclude-glob", action="append", default=None)

    # pdf backend
    ap.add_argument("--pdf-backend", choices=["auto","fitz","pdfminer","pypdf"], default="fitz")

    # progress & lowmem
    ap.add_argument("--progress", action="store_true", default=True)
    ap.add_argument("--lowmem", action="store_true", help="Single-threaded math + conservative caps")

    # output mode
    ap.add_argument("--output-mode", choices=["single","sharded"], default="sharded")

    # unknown discovery
    ap.add_argument("--discover-unknown", action="store_true", help="Collect tokens not covered by vocab")
    ap.add_argument("--min-unknown-len", type=int, default=5)
    ap.add_argument("--min-unknown-freq", type=int, default=3)
    ap.add_argument("--unknown-topk", type=int, default=20000)

    args = ap.parse_args(argv)

    input_p = Path(args.input_path).resolve()
    out_dir = Path(args.out_dir).resolve()
    vpath   = Path(args.vocab_path).resolve()
    ensure_dir(out_dir)

    # keep numeric libs single-threaded for stability/RAM
    if args.lowmem:
        os.environ.setdefault("OMP_NUM_THREADS","1")
        os.environ.setdefault("MKL_NUM_THREADS","1")
        os.environ.setdefault("OPENBLAS_NUM_THREADS","1")
        args.max_chars = min(args.max_chars, 150_000)
        args.csv_chunksize = min(args.csv_chunksize, 4000)
        args.parquet_batch_rows = min(args.parquet_batch_rows, 8192)
        if args.order == "name": args.order = "size_asc"
        if not args.max_files: args.max_files = 1000

    print(f"[INFO] Loading vocab: {vpath}")
    vocab = load_vocab(vpath, max_terms_per_group=400)
    vocab_keys = set(vocab.a2n.keys())

    # discover files (dir or single)
    if input_p.is_file():
        files = [input_p]; root = input_p.parent
    else:
        root = input_p; files = []
        for p in root.rglob("*"):
            if not p.is_file(): continue
            if p.suffix.lower() not in {".txt",".md",".rtf",".text",".pdf",".csv",".tsv",".parquet",".xlsx",".xls"}:
                continue
            files.append(p)

    # include/exclude
    if args.include_glob:
        files = [f for f in files if any(fnmatch.fnmatch(str(f).replace("\\","/"), g) for g in args.include_glob)]
    if args.exclude_glob:
        files = [f for f in files if not any(fnmatch.fnmatch(str(f).replace("\\","/"), g) for g in args.exclude_glob)]

    # order & limit
    if args.order == "name":
        files.sort(key=lambda p: str(p).lower())
    elif args.order == "size_asc":
        files.sort(key=lambda p: (p.stat().st_size if p.exists() else 0))
    else:
        files.sort(key=lambda p: -(p.stat().st_size if p.exists() else 0))
    total = len(files)
    if args.max_files and args.max_files>0: files = files[:args.max_files]
    if not files:
        print("[PRE-FLIGHT] No files found."); sys.exit(2)

    # outputs
    idx_csv = out_dir / "doc_index.csv"
    ensure_dir(idx_csv.parent)
    if not idx_csv.exists():
        with open(idx_csv, "w", encoding="utf-8", newline="") as fh:
            csv.writer(fh).writerow(["doc_path","ext","size_bytes","pages","sentences","mentions"])
    err_log = out_dir / "logs" / "errors.log"
    unk_counts_path = out_dir / "unknown_counts.jsonl"
    if args.discover_unknown and unk_counts_path.exists():
        unk_counts_path.unlink()

    outer = tqdm(total=len(files), desc="Files", unit="file", dynamic_ncols=True) if (args.progress and tqdm) else None
    totals = [0,0,0]
    for i, fp in enumerate(files, 1):
        rel = safe_rel(root, fp); size = fp.stat().st_size if fp.exists() else -1
        try:
            pages,sents,ments = process_file(
                fp, root, out_dir, vocab, vocab_keys,
                max_chars=args.max_chars, ctx=args.context_chars,
                emit_sent=args.emit_sent_text, emit_neighbors=args.emit_neighbors,
                pdf_backend=args.pdf_backend, max_pages=args.max_pages,
                max_rows=args.max_rows_per_file, csv_chunksize=args.csv_chunksize,
                parquet_batch_rows=args.parquet_batch_rows,
                progress_inner=bool(args.progress and tqdm),
                output_mode=args.output_mode,
                discover_unknown=args.discover_unknown, min_unknown_len=args.min_unknown_len,
                unk_counts_path=unk_counts_path
            )
            with open(idx_csv, "a", encoding="utf-8", newline="") as fh:
                csv.writer(fh).writerow([rel, fp.suffix.lower(), size, pages, sents, ments])
            totals[0]+=pages; totals[1]+=sents; totals[2]+=ments
            if outer: outer.update(1); outer.set_postfix(pages=pages, mentions=ments)
        except Exception as e:
            ensure_dir(err_log.parent)
            with open(err_log, "a", encoding="utf-8") as lg:
                lg.write(f"{fp}\t{type(e).__name__}: {e}\n")
                lg.write(traceback.format_exc()+"\n")
            with open(idx_csv, "a", encoding="utf-8", newline="") as fh:
                csv.writer(fh).writerow([rel, fp.suffix.lower(), size, 0, 0, 0])
            if outer: outer.update(1); outer.set_postfix(error="1")
    if outer: outer.close()

    # aggregate unknowns (bounded, pruned)
    if args.discover_unknown and unk_counts_path.exists():
        gc = collections.Counter(); prune_every = 200
        with open(unk_counts_path, "r", encoding="utf-8") as fh:
            for i, line in enumerate(fh, 1):
                rec = json.loads(line)
                for term,count in rec.get("counts", []):
                    if count >= args.min_unknown_freq: gc[term] += int(count)
                if i % prune_every == 0 and len(gc) > 2*args.unknown_topk:
                    for t,c in list(gc.items()):
                        if c < args.min_unknown_freq: del gc[t]
        top = gc.most_common(args.unknown_topk)
        out_csv = out_dir / "unknown_terms_topk.csv"
        with open(out_csv, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f); w.writerow(["term","count"]); w.writerows(top)
        print(f"[OUT] {out_csv}")

    print(f"[DONE] Docs={len(files)}/{total} | Pages={totals[0]} | Sentences={totals[1]} | Mentions={totals[2]}")
    if args.output_mode == "single":
        print(f"[OUT] {out_dir/'annotations.jsonl'}")
    else:
        print(f"[OUT] {out_dir/'docs'} (sharded)")
    print(f"[OUT] {idx_csv}")

if __name__ == "__main__":
    main()
